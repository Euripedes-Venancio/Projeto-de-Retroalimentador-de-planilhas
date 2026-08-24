import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy
import os
import xml.etree.ElementTree as ET


class Transferidor:
    # Colunas extraídas de cada item de um XML de NF-e (uma linha por <det>)
    COLUNAS_NFE = [
        "Chave NFe", "Número NF", "Data Emissão", "Arquivo XML", "CNPJ Emitente",
        "Razão Social Emitente", "Nome Fantasia Emitente", "IE Emitente", "CRT Emitente",
        "Logradouro Emitente", "Número Emitente", "Bairro Emitente",
        "Código Município Emitente", "Município Emitente", "UF Emitente", "CEP Emitente",
        "CNPJ Destinatário", "CPF Destinatário", "Razão Social Destinatário",
        "IE Destinatário", "Indicador IE Destinatário", "Email Destinatário",
        "Logradouro Destinatário", "Número Destinatário", "Bairro Destinatário",
        "Código Município Destinatário", "Município Destinatário", "UF Destinatário",
        "CEP Destinatário", "Produto Total", "Valor Desconto", "Número Item",
        "Código Produto", "Descrição Produto", "NCM", "CEST", "CFOP", "Unidade Comercial",
        "Quantidade Comercial", "Valor Unitário Comercial", "Sequência Item",
        "Total de Items", "Origem ICMS", "CST ICMS", "Base ICMS", "Alíquota ICMS",
        "Valor ICMS", "ICMS Desonerição", "Modalidade BC ICMS", "Modalidade BC ST",
        "MVAICMS ST", "Base ICMS ST", "Alíquota ICMS ST", "Valor ICMS ST", "CST IPI",
        "Base IPI", "Alíquota IPI", "Valor IPI", "CST PIS", "Base PIS", "Alíquota PIS",
        "Valor PIS", "CST COFINS", "Base COFINS", "Alíquota COFINS", "Valor COFINS",
        "Base ICMS Total", "ICMS Total", "Base ST Total", "ST Total", "IPI Total",
        "PIS Total", "COFINS Total", "Valor NF", "Modalidade Frete",
        "CNPJ Transportadora", "Nome Transportadora", "Quantidade Volumes", "Espécie",
        "Marca", "Número Volume", "Peso Líquido", "Peso Bruto", "Tipo Pagamento",
        "Valor Pagamento", "Número Protocolo", "Data Recebimento Protocolo",
        "Status Protocolo", "Motivo Protocolo", "UF", "Código NF", "Natureza Operação",
        "Modelo", "Série", "Tipo NF", "ID Destino", "Município", "Tipo Impressão",
        "Tipo Emissão", "Dígito Verificador", "Ambiente", "Finalidade",
        "Consumidor Final", "Indicador Presença", "Versão Processo",
        "Data Saída/Entrada", "ICMS Desonerição Total", "FCP Total", "FCP ST Total",
        "Frete", "Seguro", "II Total", "Outros Total", "Valor Total Tributos",
        "Informações Complementares", "EAN Produto", "EAN Tributável",
        "Unidade Tributária", "Quantidade Tributária", "Valor Unitário Tributário",
        "Indicador Total", "Informação Adicional Produto",
    ]

    def __init__(self, root):
        self.root = root
        self.root.title("Transferidor de Dados - XLSX / CSV / XML (NF-e)")
        self.root.geometry("750x640")
        self.root.resizable(False, False)
        self.root.configure(bg="#f0f4f8")

        self.df_origem = None
        self.df_destino = None          # usado só para CSV destino
        self.arquivo_destino_path = None
        self.destino_ext = None
        self.col_map = {}
        self.cols_destino = []          # nomes das colunas do destino

        self._build_ui()

    # ─────────────────────────────── UI ───────────────────────────────────────

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Arial", 10), padding=6)
        style.configure("TLabel", background="#f0f4f8", font=("Arial", 10))
        style.configure("Header.TLabel", background="#f0f4f8", font=("Arial", 13, "bold"))
        style.configure("TFrame", background="#f0f4f8")
        style.configure("Card.TFrame", background="#ffffff", relief="solid", borderwidth=1)

        ttk.Label(self.root, text="Transferidor de Dados entre Arquivos", style="Header.TLabel").pack(pady=(18, 4))
        ttk.Label(self.root, text="Suporta .xlsx, .csv e .xml (NF-e) · Preserva formatação e tabelas do destino", foreground="#6b7280").pack(pady=(0, 14))

        # Card Origem
        co = ttk.Frame(self.root, style="Card.TFrame", padding=14)
        co.grid_columnconfigure(0, weight=1)
        co.pack(fill="x", padx=24, pady=6)
        ttk.Label(co, text="📂  Arquivo de Origem", font=("Arial", 11, "bold"), background="#ffffff").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        self.entry_origem = ttk.Entry(co, width=10, state="readonly")
        self.entry_origem.grid(row=1, column=0, padx=(0, 8), sticky="ew")
        ttk.Button(co, text="Selecionar", command=self._sel_origem).grid(row=1, column=1)
        ttk.Button(co, text="↺ Limpar", command=self._limpar_origem).grid(row=1, column=2, padx=(6, 0))
        self.lbl_info_orig = ttk.Label(co, text="", foreground="#6b7280", background="#ffffff")
        self.lbl_info_orig.grid(row=2, column=0, columnspan=3, sticky="w", pady=(6, 0))

        # Card Destino
        cd = ttk.Frame(self.root, style="Card.TFrame", padding=14)
        cd.grid_columnconfigure(0, weight=1)
        cd.pack(fill="x", padx=24, pady=6)
        ttk.Label(cd, text="💾  Arquivo de Destino", font=("Arial", 11, "bold"), background="#ffffff").grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        self.entry_destino = ttk.Entry(cd, width=52, state="readonly")
        self.entry_destino.grid(row=1, column=0, padx=(0, 8), sticky="ew")
        ttk.Button(cd, text="Selecionar", command=self._sel_destino).grid(row=1, column=1)
        ttk.Button(cd, text="↺ Limpar", command=self._limpar_destino).grid(row=1, column=2, padx=(6, 0))
        self.lbl_info_dest = ttk.Label(cd, text="", foreground="#6b7280", background="#ffffff")
        self.lbl_info_dest.grid(row=2, column=0, columnspan=3, sticky="w", pady=(6, 0))

        # Card Mapeamento
        cm = ttk.Frame(self.root, style="Card.TFrame", padding=14)
        cm.pack(fill="x", padx=24, pady=6)
        ttk.Label(cm, text="🔗  Mapeamento de Colunas", font=("Arial", 11, "bold"), background="#ffffff").pack(anchor="w", pady=(0, 8))
        self.frame_mapa = ttk.Frame(cm, style="Card.TFrame")
        self.frame_mapa.pack(fill="x")
        self.lbl_mapa_hint = ttk.Label(cm, text="Selecione os dois arquivos para configurar o mapeamento.", foreground="#9ca3af", background="#ffffff")
        self.lbl_mapa_hint.pack(anchor="w")

        # Botões
        bf = ttk.Frame(self.root, style="TFrame")
        bf.pack(pady=14)
        self.btn_transferir = ttk.Button(bf, text="▶  Transferir Dados", command=self._transferir, state="disabled")
        self.btn_transferir.grid(row=0, column=0, padx=8)
        ttk.Button(bf, text="📋  Pré-visualizar Origem", command=self._preview).grid(row=0, column=1, padx=8)

        # Log
        lf = ttk.Frame(self.root, style="Card.TFrame", padding=8)
        lf.pack(fill="both", expand=True, padx=24, pady=(0, 18))
        ttk.Label(lf, text="Log", font=("Arial", 9, "bold"), background="#ffffff").pack(anchor="w")
        self.log = tk.Text(lf, height=6, state="disabled", wrap="word",
                           font=("Courier", 9), bg="#f9fafb", relief="flat", bd=0)
        self.log.pack(fill="both", expand=True)

    def _log(self, msg):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.configure(state="disabled")
        self.log.see("end")

    def _set_entry(self, entry, valor):
        entry.configure(state="normal")
        entry.delete(0, "end")
        entry.insert(0, valor)
        entry.configure(state="readonly")

    # ─────────────────────── leitura de arquivos ───────────────────────────────

    def _ler_df(self, path, duplo_cabecalho=True):
        """
        duplo_cabecalho=True → ignora linha 1, usa linha 2 como cabeçalho,
        lê dados a partir da linha 3.
        """
        ext = os.path.splitext(path)[1].lower()
        header_row = 1 if duplo_cabecalho else 0   # 0-based para pandas
        if ext in (".xlsx", ".xlsm", ".xltx"):
            return pd.read_excel(path, header=header_row, dtype=str)
        elif ext == ".csv":
            with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
                amostra = f.read(2048)
            sep = ";" if amostra.count(";") > amostra.count(",") else ","
            return pd.read_csv(path, sep=sep, header=header_row, dtype=str, encoding="utf-8-sig")
        else:
            raise ValueError(f"Formato não suportado: {ext}")

    def _colunas_destino_xlsx(self, path):
        """Lê cabeçalhos da linha 2 do xlsx de destino (duplo cabeçalho)."""
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        # linha 2 = cabeçalho real das colunas de dados
        for cell in next(ws.iter_rows(min_row=2, max_row=2)):
            v = cell.value
            if v is not None:
                headers.append(str(v))
        wb.close()
        return headers

    # ─────────────────────── leitura de XML (NF-e) ─────────────────────────────

    @staticmethod
    def _strip_ns(elem):
        """Remove o namespace (ex: {http://www.portalfiscal.inf.br/nfe}) de todas as tags,
        para permitir usar find()/findall() sem prefixo."""
        for el in elem.iter():
            if "}" in el.tag:
                el.tag = el.tag.split("}", 1)[1]
        return elem

    @staticmethod
    def _txt(elem, path, default=""):
        if elem is None:
            return default
        found = elem.find(path)
        if found is not None and found.text is not None:
            return found.text
        return default

    def _ler_xml_nfe(self, path):
        """Lê um XML de NF-e (nfeProc ou NFe avulso) e retorna uma lista de dicts,
        uma linha por item (<det>), com as colunas definidas em COLUNAS_NFE."""
        tree = ET.parse(path)
        root = self._strip_ns(tree.getroot())

        if root.tag == "nfeProc":
            nfe = root.find("NFe")
            protNFe = root.find("protNFe")
        elif root.tag == "NFe":
            nfe = root
            protNFe = None
        else:
            raise ValueError(f"XML não reconhecido como NF-e (tag raiz '{root.tag}'): {os.path.basename(path)}")

        if nfe is None:
            raise ValueError(f"Não foi possível localizar a tag <NFe> em: {os.path.basename(path)}")

        infNFe = nfe.find("infNFe")
        ide = infNFe.find("ide")
        emit = infNFe.find("emit")
        enderEmit = emit.find("enderEmit") if emit is not None else None
        dest = infNFe.find("dest")
        enderDest = dest.find("enderDest") if dest is not None else None
        total = infNFe.find("total")
        icmsTot = total.find("ICMSTot") if total is not None else None
        transp = infNFe.find("transp")
        transporta = transp.find("transporta") if transp is not None else None
        vol = transp.find("vol") if transp is not None else None
        pag = infNFe.find("pag")
        detPag = pag.find("detPag") if pag is not None else None
        infAdic = infNFe.find("infAdic")
        infProt = protNFe.find("infProt") if protNFe is not None else None

        chave = infNFe.get("Id", "")
        if chave.startswith("NFe"):
            chave = chave[3:]

        dets = infNFe.findall("det")
        total_itens = len(dets)
        nome_arquivo = os.path.basename(path)

        linhas = []
        for i, det in enumerate(dets, start=1):
            prod = det.find("prod")
            imposto = det.find("imposto")
            icms_elem = imposto.find("ICMS") if imposto is not None else None
            icms_sub = list(icms_elem)[0] if icms_elem is not None and len(icms_elem) else None
            ipi_elem = imposto.find("IPI") if imposto is not None else None
            ipi_sub = list(ipi_elem)[0] if ipi_elem is not None and len(ipi_elem) else None
            pis_elem = imposto.find("PIS") if imposto is not None else None
            pis_sub = list(pis_elem)[0] if pis_elem is not None and len(pis_elem) else None
            cofins_elem = imposto.find("COFINS") if imposto is not None else None
            cofins_sub = list(cofins_elem)[0] if cofins_elem is not None and len(cofins_elem) else None

            linhas.append({
                "Chave NFe": chave,
                "Número NF": self._txt(ide, "nNF"),
                "Data Emissão": self._txt(ide, "dhEmi"),
                "Arquivo XML": nome_arquivo,
                "CNPJ Emitente": self._txt(emit, "CNPJ"),
                "Razão Social Emitente": self._txt(emit, "xNome"),
                "Nome Fantasia Emitente": self._txt(emit, "xFant"),
                "IE Emitente": self._txt(emit, "IE"),
                "CRT Emitente": self._txt(emit, "CRT"),
                "Logradouro Emitente": self._txt(enderEmit, "xLgr"),
                "Número Emitente": self._txt(enderEmit, "nro"),
                "Bairro Emitente": self._txt(enderEmit, "xBairro"),
                "Código Município Emitente": self._txt(enderEmit, "cMun"),
                "Município Emitente": self._txt(enderEmit, "xMun"),
                "UF Emitente": self._txt(enderEmit, "UF"),
                "CEP Emitente": self._txt(enderEmit, "CEP"),
                "CNPJ Destinatário": self._txt(dest, "CNPJ"),
                "CPF Destinatário": self._txt(dest, "CPF"),
                "Razão Social Destinatário": self._txt(dest, "xNome"),
                "IE Destinatário": self._txt(dest, "IE"),
                "Indicador IE Destinatário": self._txt(dest, "indIEDest"),
                "Email Destinatário": self._txt(dest, "email"),
                "Logradouro Destinatário": self._txt(enderDest, "xLgr"),
                "Número Destinatário": self._txt(enderDest, "nro"),
                "Bairro Destinatário": self._txt(enderDest, "xBairro"),
                "Código Município Destinatário": self._txt(enderDest, "cMun"),
                "Município Destinatário": self._txt(enderDest, "xMun"),
                "UF Destinatário": self._txt(enderDest, "UF"),
                "CEP Destinatário": self._txt(enderDest, "CEP"),
                "Produto Total": self._txt(prod, "vProd"),
                "Valor Desconto": self._txt(prod, "vDesc"),
                "Número Item": det.get("nItem", ""),
                "Código Produto": self._txt(prod, "cProd"),
                "Descrição Produto": self._txt(prod, "xProd"),
                "NCM": self._txt(prod, "NCM"),
                "CEST": self._txt(prod, "CEST"),
                "CFOP": self._txt(prod, "CFOP"),
                "Unidade Comercial": self._txt(prod, "uCom"),
                "Quantidade Comercial": self._txt(prod, "qCom"),
                "Valor Unitário Comercial": self._txt(prod, "vUnCom"),
                "Sequência Item": str(i),
                "Total de Items": str(total_itens),
                "Origem ICMS": self._txt(icms_sub, "orig"),
                "CST ICMS": self._txt(icms_sub, "CST") or self._txt(icms_sub, "CSOSN"),
                "Base ICMS": self._txt(icms_sub, "vBC"),
                "Alíquota ICMS": self._txt(icms_sub, "pICMS"),
                "Valor ICMS": self._txt(icms_sub, "vICMS"),
                "ICMS Desonerição": self._txt(icms_sub, "vICMSDeson"),
                "Modalidade BC ICMS": self._txt(icms_sub, "modBC"),
                "Modalidade BC ST": self._txt(icms_sub, "modBCST"),
                "MVAICMS ST": self._txt(icms_sub, "pMVAST"),
                "Base ICMS ST": self._txt(icms_sub, "vBCST"),
                "Alíquota ICMS ST": self._txt(icms_sub, "pICMSST"),
                "Valor ICMS ST": self._txt(icms_sub, "vICMSST"),
                "CST IPI": self._txt(ipi_sub, "CST"),
                "Base IPI": self._txt(ipi_sub, "vBC"),
                "Alíquota IPI": self._txt(ipi_sub, "pIPI"),
                "Valor IPI": self._txt(ipi_sub, "vIPI"),
                "CST PIS": self._txt(pis_sub, "CST"),
                "Base PIS": self._txt(pis_sub, "vBC"),
                "Alíquota PIS": self._txt(pis_sub, "pPIS"),
                "Valor PIS": self._txt(pis_sub, "vPIS"),
                "CST COFINS": self._txt(cofins_sub, "CST"),
                "Base COFINS": self._txt(cofins_sub, "vBC"),
                "Alíquota COFINS": self._txt(cofins_sub, "pCOFINS"),
                "Valor COFINS": self._txt(cofins_sub, "vCOFINS"),
                "Base ICMS Total": self._txt(icmsTot, "vBC"),
                "ICMS Total": self._txt(icmsTot, "vICMS"),
                "Base ST Total": self._txt(icmsTot, "vBCST"),
                "ST Total": self._txt(icmsTot, "vST"),
                "IPI Total": self._txt(icmsTot, "vIPI"),
                "PIS Total": self._txt(icmsTot, "vPIS"),
                "COFINS Total": self._txt(icmsTot, "vCOFINS"),
                "Valor NF": self._txt(icmsTot, "vNF"),
                "Modalidade Frete": self._txt(transp, "modFrete"),
                "CNPJ Transportadora": self._txt(transporta, "CNPJ"),
                "Nome Transportadora": self._txt(transporta, "xNome"),
                "Quantidade Volumes": self._txt(vol, "qVol"),
                "Espécie": self._txt(vol, "esp"),
                "Marca": self._txt(vol, "marca"),
                "Número Volume": self._txt(vol, "nVol"),
                "Peso Líquido": self._txt(vol, "pesoL"),
                "Peso Bruto": self._txt(vol, "pesoB"),
                "Tipo Pagamento": self._txt(detPag, "tPag"),
                "Valor Pagamento": self._txt(detPag, "vPag"),
                "Número Protocolo": self._txt(infProt, "nProt"),
                "Data Recebimento Protocolo": self._txt(infProt, "dhRecbto"),
                "Status Protocolo": self._txt(infProt, "cStat"),
                "Motivo Protocolo": self._txt(infProt, "xMotivo"),
                "UF": self._txt(ide, "cUF"),
                "Código NF": self._txt(ide, "cNF"),
                "Natureza Operação": self._txt(ide, "natOp"),
                "Modelo": self._txt(ide, "mod"),
                "Série": self._txt(ide, "serie"),
                "Tipo NF": self._txt(ide, "tpNF"),
                "ID Destino": self._txt(ide, "idDest"),
                "Município": self._txt(ide, "cMunFG"),
                "Tipo Impressão": self._txt(ide, "tpImp"),
                "Tipo Emissão": self._txt(ide, "tpEmis"),
                "Dígito Verificador": self._txt(ide, "cDV"),
                "Ambiente": self._txt(ide, "tpAmb"),
                "Finalidade": self._txt(ide, "finNFe"),
                "Consumidor Final": self._txt(ide, "indFinal"),
                "Indicador Presença": self._txt(ide, "indPres"),
                "Versão Processo": self._txt(ide, "verProc"),
                "Data Saída/Entrada": self._txt(ide, "dhSaiEnt"),
                "ICMS Desonerição Total": self._txt(icmsTot, "vICMSDeson"),
                "FCP Total": self._txt(icmsTot, "vFCP"),
                "FCP ST Total": self._txt(icmsTot, "vFCPST"),
                "Frete": self._txt(icmsTot, "vFrete"),
                "Seguro": self._txt(icmsTot, "vSeg"),
                "II Total": self._txt(icmsTot, "vII"),
                "Outros Total": self._txt(icmsTot, "vOutro"),
                "Valor Total Tributos": self._txt(icmsTot, "vTotTrib"),
                "Informações Complementares": self._txt(infAdic, "infCpl"),
                "EAN Produto": self._txt(prod, "cEAN"),
                "EAN Tributável": self._txt(prod, "cEANTrib"),
                "Unidade Tributária": self._txt(prod, "uTrib"),
                "Quantidade Tributária": self._txt(prod, "qTrib"),
                "Valor Unitário Tributário": self._txt(prod, "vUnTrib"),
                "Indicador Total": self._txt(prod, "indTot"),
                "Informação Adicional Produto": self._txt(det, "infAdProd"),
            })
        return linhas

    def _ler_xml_multiplos(self, paths):
        """Lê vários arquivos XML de NF-e e concatena todas as linhas de itens."""
        linhas = []
        erros = []
        for p in paths:
            try:
                linhas.extend(self._ler_xml_nfe(p))
            except Exception as e:
                erros.append(f"{os.path.basename(p)}: {e}")
        if erros:
            self._log("⚠ Arquivos XML com erro (ignorados):\n  " + "\n  ".join(erros))
        if not linhas:
            raise ValueError("Nenhum item de NF-e foi extraído dos XML selecionados.")
        return pd.DataFrame(linhas, columns=self.COLUNAS_NFE, dtype=str)

    # ─────────────────────── seleção / limpeza ────────────────────────────────

    def _sel_origem(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar arquivo(s) de origem (xlsx/csv: apenas 1 · xml: pode selecionar vários)",
            filetypes=[
                ("Todos suportados", "*.xlsx *.xlsm *.xltx *.csv *.xml"),
                ("Planilhas e CSV", "*.xlsx *.xlsm *.xltx *.csv"),
                ("XML NF-e", "*.xml"),
                ("Todos", "*.*"),
            ])
        if not paths:
            return
        try:
            ext_primeiro = os.path.splitext(paths[0])[1].lower()

            if ext_primeiro == ".xml":
                self.df_origem = self._ler_xml_multiplos(paths)
                label = f"{len(paths)} arquivo(s) XML" if len(paths) > 1 else paths[0]
                origem_log = f"{len(paths)} XML(s) de NF-e"
            else:
                if len(paths) > 1:
                    self._log("⚠ Vários arquivos selecionados; para xlsx/csv apenas o primeiro é usado.")
                path = paths[0]
                self.df_origem = self._ler_df(path)
                label = path
                origem_log = os.path.basename(path)

            self._set_entry(self.entry_origem, label)
            nr, nc = self.df_origem.shape
            self.lbl_info_orig.configure(text=f"✔  {nr} linhas · {nc} colunas: {', '.join(self.df_origem.columns)}")
            self._log(f"Origem: {origem_log}  ({nr} linhas, {nc} colunas)")
            self._atualizar_mapeamento()
        except Exception as e:
            messagebox.showerror("Erro ao abrir origem", str(e))

    def _sel_destino(self):
        path = filedialog.askopenfilename(title="Selecionar arquivo de destino",
            filetypes=[("Planilhas e CSV", "*.xlsx *.xlsm *.xltx *.csv"), ("Todos", "*.*")])
        if not path:
            return
        try:
            ext = os.path.splitext(path)[1].lower()
            self.arquivo_destino_path = path
            self.destino_ext = ext

            if ext in (".xlsx", ".xlsm", ".xltx"):
                self.cols_destino = self._colunas_destino_xlsx(path)
                self.df_destino = None
                nr_info = "xlsx — formatação e tabelas serão preservadas"
            else:
                self.df_destino = self._ler_df(path, duplo_cabecalho=True)
                self.cols_destino = list(self.df_destino.columns)
                nr_info = f"{len(self.df_destino)} linhas existentes"

            self._set_entry(self.entry_destino, path)
            self.lbl_info_dest.configure(
                text=f"✔  {len(self.cols_destino)} colunas · {nr_info}: {', '.join(self.cols_destino)}")
            self._log(f"Destino: {os.path.basename(path)}  ({len(self.cols_destino)} colunas)")
            self._atualizar_mapeamento()
        except Exception as e:
            messagebox.showerror("Erro ao abrir destino", str(e))

    def _limpar_origem(self):
        self.df_origem = None
        self._set_entry(self.entry_origem, "")
        self.lbl_info_orig.configure(text="")
        self._atualizar_mapeamento()

    def _limpar_destino(self):
        self.df_destino = None
        self.arquivo_destino_path = None
        self.destino_ext = None
        self.cols_destino = []
        self._set_entry(self.entry_destino, "")
        self.lbl_info_dest.configure(text="")
        self._atualizar_mapeamento()

    # ─────────────────────── mapeamento de colunas ────────────────────────────

    def _atualizar_mapeamento(self):
        for w in self.frame_mapa.winfo_children():
            w.destroy()
        self.col_map = {}

        if self.df_origem is None or not self.cols_destino:
            self.lbl_mapa_hint.configure(text="Selecione os dois arquivos para configurar o mapeamento.")
            self.btn_transferir.configure(state="disabled")
            return

        self.lbl_mapa_hint.configure(text="")
        opcoes = ["-- ignorar --"] + self.cols_destino

        ttk.Label(self.frame_mapa, text="Coluna na Origem", font=("Arial", 9, "bold"),
                  background="#ffffff", width=28).grid(row=0, column=0, sticky="w")
        ttk.Label(self.frame_mapa, text="→  Coluna no Destino", font=("Arial", 9, "bold"),
                  background="#ffffff", width=30).grid(row=0, column=1, sticky="w")

        canvas = tk.Canvas(self.frame_mapa,
                           height=min(len(self.df_origem.columns) * 28 + 10, 130),
                           bg="#ffffff", highlightthickness=0)
        scroll = ttk.Scrollbar(self.frame_mapa, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas, style="Card.TFrame")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.grid(row=1, column=0, columnspan=2, sticky="ew")
        scroll.grid(row=1, column=2, sticky="ns")

        for i, col in enumerate(self.df_origem.columns):
            ttk.Label(inner, text=col, background="#ffffff", width=28).grid(row=i, column=0, sticky="w", pady=1)
            var = tk.StringVar(value=col if col in self.cols_destino else "-- ignorar --")
            ttk.Combobox(inner, textvariable=var, values=opcoes, state="readonly", width=26).grid(
                row=i, column=1, sticky="w", padx=(8, 0), pady=1)
            self.col_map[col] = var

        self.btn_transferir.configure(state="normal")

    # ─────────────────────── pré-visualização ─────────────────────────────────

    def _preview(self):
        if self.df_origem is None:
            messagebox.showinfo("Aviso", "Selecione o arquivo de origem primeiro.")
            return
        win = tk.Toplevel(self.root)
        win.title("Pré-visualização — Origem (primeiras 50 linhas)")
        win.geometry("800x400")
        f = ttk.Frame(win)
        f.pack(fill="both", expand=True, padx=8, pady=8)
        cols = list(self.df_origem.columns)
        tree = ttk.Treeview(f, columns=cols, show="headings")
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=max(80, min(180, len(c) * 11)))
        for _, row in self.df_origem.head(50).iterrows():
            tree.insert("", "end", values=list(row))
        sv = ttk.Scrollbar(f, orient="vertical", command=tree.yview)
        sh = ttk.Scrollbar(f, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=sv.set, xscrollcommand=sh.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sv.grid(row=0, column=1, sticky="ns")
        sh.grid(row=1, column=0, sticky="ew")
        f.rowconfigure(0, weight=1)
        f.columnconfigure(0, weight=1)

    # ─────────────────────── transferência ────────────────────────────────────

    def _transferir(self):
        if self.df_origem is None or not self.arquivo_destino_path:
            messagebox.showwarning("Atenção", "Selecione os dois arquivos antes de transferir.")
            return

        mapa = {orig: var.get() for orig, var in self.col_map.items()
                if var.get() != "-- ignorar --"}
        if not mapa:
            messagebox.showwarning("Atenção", "Nenhuma coluna mapeada para transferir.")
            return

        ext = self.destino_ext
        ft = [("Excel", "*.xlsx")] if ext in (".xlsx", ".xlsm", ".xltx") else [("CSV", "*.csv")]
        save_path = filedialog.asksaveasfilename(
            title="Salvar arquivo de destino atualizado",
            initialfile=os.path.basename(self.arquivo_destino_path),
            defaultextension=ext,
            filetypes=ft + [("Todos", "*.*")])
        if not save_path:
            return

        try:
            if ext in (".xlsx", ".xlsm", ".xltx"):
                linhas = self._transferir_xlsx(mapa, save_path)
            else:
                linhas = self._transferir_csv(mapa, save_path)

            self._log(f"✔ Concluído! {linhas} linhas adicionadas → {os.path.basename(save_path)}")
            messagebox.showinfo("Sucesso",
                f"{linhas} linhas transferidas com sucesso!\n\nSalvo em:\n{save_path}")
        except Exception as e:
            self._log(f"✘ Erro: {e}")
            messagebox.showerror("Erro na transferência", str(e))

    # ── xlsx: preserva tabela e formatação ────────────────────────────────────

    def _transferir_xlsx(self, mapa, save_path):
        import shutil

        # copia o arquivo destino para o caminho de saída (preserva tudo)
        shutil.copy2(self.arquivo_destino_path, save_path)

        wb = load_workbook(save_path)
        ws = wb.active

        # cabeçalhos estão na linha 2 (linha 1 é título/cabeçalho extra)
        LINHA_CABECALHO = 2
        PRIMEIRA_LINHA_DADOS = 3

        header_idx = {}
        for cell in ws[LINHA_CABECALHO]:
            if cell.value is not None:
                header_idx[str(cell.value)] = cell.column  # 1-based

        # detecta a última linha com dados (a partir da linha 3)
        ultima_linha = ws.max_row
        while ultima_linha >= PRIMEIRA_LINHA_DADOS:
            if any(ws.cell(row=ultima_linha, column=c).value is not None
                   for c in range(1, ws.max_column + 1)):
                break
            ultima_linha -= 1

        # se não há nenhuma linha de dados ainda, começa na linha 3
        if ultima_linha < PRIMEIRA_LINHA_DADOS:
            ultima_linha = PRIMEIRA_LINHA_DADOS - 1

        # linha-modelo para copiar estilo: usa última linha de dados,
        # ou a linha de cabeçalho se ainda não há dados
        linha_modelo = ultima_linha if ultima_linha >= PRIMEIRA_LINHA_DADOS else LINHA_CABECALHO

        linhas_adicionadas = 0
        for _, row_orig in self.df_origem.iterrows():
            nova_linha_num = ultima_linha + 1 + linhas_adicionadas

            # copia estilo da linha-modelo para a nova linha
            for col_num in range(1, ws.max_column + 1):
                cel_modelo = ws.cell(row=linha_modelo, column=col_num)
                cel_nova = ws.cell(row=nova_linha_num, column=col_num)
                if cel_modelo.has_style:
                    cel_nova.font = copy.copy(cel_modelo.font)
                    cel_nova.border = copy.copy(cel_modelo.border)
                    cel_nova.fill = copy.copy(cel_modelo.fill)
                    cel_nova.number_format = cel_modelo.number_format
                    cel_nova.alignment = copy.copy(cel_modelo.alignment)

            # preenche os valores mapeados
            for col_orig, col_dest in mapa.items():
                if col_dest in header_idx:
                    valor = row_orig.get(col_orig, None)
                    if pd.isna(valor) if isinstance(valor, float) else str(valor) == "nan":
                        valor = None
                    ws.cell(row=nova_linha_num, column=header_idx[col_dest]).value = valor

            linhas_adicionadas += 1

        # expande o intervalo da tabela (Table), se existir
        for tbl in ws.tables.values():
            ref_orig = tbl.ref  # ex: "A2:F10"
            col_ini, linha_ini, col_fim, _ = self._parse_ref(ref_orig)
            nova_ref = f"{col_ini}{linha_ini}:{col_fim}{ultima_linha + linhas_adicionadas}"
            tbl.ref = nova_ref
            self._log(f"  Tabela '{tbl.displayName}' expandida: {ref_orig} → {nova_ref}")

        wb.save(save_path)
        wb.close()
        return linhas_adicionadas

    def _parse_ref(self, ref):
        """Divide 'A1:F10' em (col_ini_letra, linha_ini, col_fim_letra, linha_fim)."""
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        inicio, fim = ref.split(":")
        col_ini_l, linha_ini = coordinate_from_string(inicio)
        col_fim_l, linha_fim = coordinate_from_string(fim)
        return col_ini_l, linha_ini, col_fim_l, linha_fim

    # ── csv: simples ──────────────────────────────────────────────────────────

    def _transferir_csv(self, mapa, save_path):
        # lê destino preservando as 2 linhas de cabeçalho como texto bruto
        with open(self.arquivo_destino_path, "r", encoding="utf-8-sig", errors="replace") as f:
            amostra = f.read(2048)
        sep = ";" if amostra.count(";") > amostra.count(",") else ","

        # lê as 2 linhas de cabeçalho brutas para reescrever no topo
        with open(self.arquivo_destino_path, "r", encoding="utf-8-sig", errors="replace") as f:
            linha_cab1 = f.readline()
            linha_cab2 = f.readline()

        # lê os dados existentes (a partir da linha 3)
        df_dest = self.df_destino.copy()  # já foi lido com header=1 (linha 2)

        novas = []
        for _, row_orig in self.df_origem.iterrows():
            nova = {c: None for c in df_dest.columns}
            for col_orig, col_dest in mapa.items():
                nova[col_dest] = row_orig.get(col_orig)
            novas.append(nova)

        df_final = pd.concat([df_dest, pd.DataFrame(novas)], ignore_index=True)

        # grava: linha1 + linha2 (cabeçalhos originais) + dados
        with open(save_path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(linha_cab1)
            f.write(linha_cab2)
            f.write(df_final.to_csv(index=False, sep=sep, lineterminator="\n",
                                    header=False))
        return len(novas)


if __name__ == "__main__":
    root = tk.Tk()
    app = Transferidor(root)
    root.mainloop()